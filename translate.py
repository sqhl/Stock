import os
import json
from openpyxl import load_workbook,Workbook

def input(path):
    wb = load_workbook(path)
    ws = wb.active
    i = 1
    fields = []
    data = []
    for row in ws.rows:
        list = []
        for cell in row:
            aa = str(cell.value)
            if (aa == ""):
                aa = "1"
            list.append(aa)
        if (i < 1):
            pass
        elif (i == 1):
            fields = list
        else:
            data.append(list[:2])
        i = i + 1
    JsonD = json.dumps(data)
    with open(os.getcwd()+"/Models/data.json", "w") as f:
        f.write(JsonD)

def output(path, D):
    wb = Workbook()
    sheet = wb.active
    sheet.append(["代码","名称"])
    for i in D:
        sheet.append(i[:2])
    wb.save(path+'/导出股票池.xlsx')

def main(path):
    # input(path)
    output(path)

if __name__ == "__main__":
    main('./A.xlsx')